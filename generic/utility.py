import openpyxl
class Excel:
    @staticmethod
    def get_data(filepath,sheet_name,row,col):
      try:
           wb= openpyxl.load_workbook(filepath)
           sheet=wb[sheet_name]
           value=sheet.cell(row,col).value
           if value==None:
               print("Cell value is NONE")
               value=""
      except Exception as e:
           print(str(e))
           value=""

      return value

    @staticmethod
    def get_rowcount(filepath,sheet_name):
      try:
           wb= openpyxl.load_workbook(filepath)
           sheet=wb[sheet_name]
           rc=sheet.max_row
      except Exception as e:
           print(str(e))
           rc=0

      return rc